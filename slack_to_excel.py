#!/usr/bin/env python3
"""
Slack Channel Backup JSON to Excel Converter
슬랙 채널 백업 JSON 파일을 엑셀 파일로 변환하는 스크립트
"""

import json
import pandas as pd
import argparse
import os
import glob
from datetime import datetime
from typing import Dict, List, Any
import re
import openpyxl
from openpyxl.styles import Alignment

def clean_text(text: str) -> str:
    """텍스트에서 HTML 태그와 특수 문자를 정리합니다."""
    if not text:
        return ""
    
    # HTML 태그 제거
    text = re.sub(r'<[^>]+>', '', text)
    # 슬랙 이모지 제거 (예: :smile:)
    text = re.sub(r':[a-zA-Z0-9_+-]+:', '', text)
    # 멀티라인 텍스트 정리 (줄바꿈 유지)
    text = text.replace('\r\n', '\n').replace('\r', '\n')
    # 연속된 공백 제거 (줄바꿈은 유지)
    text = re.sub(r'[ \t]+', ' ', text)
    # 연속된 줄바꿈을 하나로 정리
    text = re.sub(r'\n\s*\n', '\n\n', text)
    
    return text.strip()

def extract_user_info(users: List[Dict]) -> pd.DataFrame:
    """사용자 정보를 추출하여 DataFrame으로 변환합니다."""
    user_data = []
    
    for user in users:
        user_info = {
            'user_id': user.get('id', ''),
            'username': user.get('name', ''),
            'real_name': user.get('real_name', ''),
            'display_name': user.get('profile', {}).get('display_name', ''),
            'email': user.get('profile', {}).get('email', ''),
            'is_bot': user.get('is_bot', False),
            'is_admin': user.get('is_admin', False),
            'is_owner': user.get('is_owner', False),
            'deleted': user.get('deleted', False),
            'created': user.get('created', ''),
            'updated': user.get('updated', '')
        }
        user_data.append(user_info)
    
    return pd.DataFrame(user_data)

def extract_channel_info(channels: List[Dict]) -> pd.DataFrame:
    """채널 정보를 추출하여 DataFrame으로 변환합니다."""
    channel_data = []
    
    for channel in channels:
        channel_info = {
            'channel_id': channel.get('id', ''),
            'channel_name': channel.get('name', ''),
            'channel_type': channel.get('is_private', False) and 'private' or 'public',
            'topic': channel.get('topic', {}).get('value', ''),
            'purpose': channel.get('purpose', {}).get('value', ''),
            'member_count': channel.get('num_members', 0),
            'created': channel.get('created', ''),
            'creator': channel.get('creator', '')
        }
        channel_data.append(channel_info)
    
    return pd.DataFrame(channel_data)

def extract_messages(messages: List[Dict], users_dict: Dict, source_file: str = "") -> pd.DataFrame:
    """메시지 정보를 추출하여 DataFrame으로 변환합니다."""
    message_data = []
    
    for msg in messages:
        # 기본 메시지 정보
        message_info = {
            'source_file': source_file,
            'message_id': msg.get('client_msg_id', msg.get('ts', '')),
            'timestamp': msg.get('ts', ''),
            'datetime': datetime.fromtimestamp(float(msg.get('ts', 0))).strftime('%Y-%m-%d %H:%M:%S') if msg.get('ts') else '',
            'user_id': msg.get('user', ''),
            'username': users_dict.get(msg.get('user', ''), 'Unknown'),
            'text': clean_text(msg.get('text', '')),
            'type': msg.get('type', 'message'),
            'subtype': msg.get('subtype', ''),
            'thread_ts': msg.get('thread_ts', ''),
            'reply_count': msg.get('reply_count', 0),
            'reply_users_count': msg.get('reply_users_count', 0),
            'latest_reply': msg.get('latest_reply', ''),
            'reactions': '',
            'files': '',
            'attachments': ''
        }
        
        # 리액션 정보 추출
        if msg.get('reactions'):
            reactions = []
            for reaction in msg['reactions']:
                reactions.append(f"{reaction.get('name', '')}({reaction.get('count', 0)})")
            message_info['reactions'] = ', '.join(reactions)
        
        # 파일 정보 추출
        if msg.get('files'):
            files = []
            for file in msg['files']:
                files.append(f"{file.get('name', '')} ({file.get('filetype', '')})")
            message_info['files'] = ', '.join(files)
        
        # 첨부파일 정보 추출
        if msg.get('attachments'):
            attachments = []
            for attachment in msg['attachments']:
                attachments.append(attachment.get('title', attachment.get('text', ''))[:50])
            message_info['attachments'] = ', '.join(attachments)
        
        message_data.append(message_info)
    
    return pd.DataFrame(message_data)

def process_single_json_file(json_file_path: str, users_dict: Dict = None) -> pd.DataFrame:
    """단일 JSON 파일을 처리하여 메시지 DataFrame을 반환합니다."""
    print(f"  - {os.path.basename(json_file_path)} 처리 중...")
    
    try:
        with open(json_file_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        # 데이터가 리스트인 경우 (메시지 배열)
        if isinstance(data, list):
            return extract_messages(data, users_dict or {}, os.path.basename(json_file_path))
        
        # 데이터가 딕셔너리인 경우 (전체 백업 구조)
        elif isinstance(data, dict):
            if 'messages' in data:
                return extract_messages(data['messages'], users_dict or {}, os.path.basename(json_file_path))
            else:
                print(f"    경고: {json_file_path}에 메시지 데이터가 없습니다.")
                return pd.DataFrame()
        
        else:
            print(f"    경고: {json_file_path}의 데이터 형식을 인식할 수 없습니다.")
            return pd.DataFrame()
            
    except Exception as e:
        print(f"    오류: {json_file_path} 처리 중 오류 발생 - {str(e)}")
        return pd.DataFrame()

def convert_slack_json_to_excel(json_file_path: str, output_file_path: str = None):
    """슬랙 JSON 백업 파일을 엑셀 파일로 변환합니다."""
    
    # JSON 파일 읽기
    print(f"JSON 파일을 읽는 중: {json_file_path}")
    with open(json_file_path, 'r', encoding='utf-8') as f:
        data = json.load(f)
    
    # 출력 파일명 설정
    if not output_file_path:
        base_name = os.path.splitext(json_file_path)[0]
        output_file_path = f"{base_name}_converted.xlsx"
    
    print(f"엑셀 파일로 변환 중: {output_file_path}")
    
    # ExcelWriter 객체 생성
    with pd.ExcelWriter(output_file_path, engine='openpyxl') as writer:
        
        # 사용자 정보 처리
        if 'users' in data:
            print("사용자 정보 처리 중...")
            users_df = extract_user_info(data['users'])
            users_df.to_excel(writer, sheet_name='Users', index=False)
            
            # 사용자 ID와 이름 매핑 생성
            users_dict = {user['id']: user.get('real_name', user.get('name', 'Unknown')) 
                         for user in data['users']}
        else:
            users_dict = {}
        
        # 채널 정보 처리
        if 'channels' in data:
            print("채널 정보 처리 중...")
            channels_df = extract_channel_info(data['channels'])
            channels_df.to_excel(writer, sheet_name='Channels', index=False)
        
        # 메시지 정보 처리
        if 'messages' in data:
            print("메시지 정보 처리 중...")
            messages_df = extract_messages(data['messages'], users_dict)
            messages_df.to_excel(writer, sheet_name='Messages', index=False)
            
            # 줄바꿈 표시를 위한 셀 서식 설정
            worksheet = writer.sheets['Messages']
            for row in worksheet.iter_rows(min_row=2, min_col=6, max_col=6):  # text 컬럼
                for cell in row:
                    if cell.value and '\n' in str(cell.value):
                        cell.alignment = openpyxl.styles.Alignment(wrap_text=True)
        
        # 개별 채널별 메시지 처리
        if 'channels' in data:
            print("채널별 메시지 처리 중...")
            for channel in data['channels']:
                channel_name = channel.get('name', channel.get('id', 'unknown'))
                if 'messages' in channel:
                    print(f"  - {channel_name} 채널 처리 중...")
                    channel_messages_df = extract_messages(channel['messages'], users_dict)
                    # 시트명 길이 제한 (Excel 시트명은 31자 제한)
                    sheet_name = channel_name[:31] if len(channel_name) <= 31 else channel_name[:28] + '...'
                    channel_messages_df.to_excel(writer, sheet_name=sheet_name, index=False)
                    
                    # 줄바꿈 표시를 위한 셀 서식 설정
                    worksheet = writer.sheets[sheet_name]
                    for row in worksheet.iter_rows(min_row=2, min_col=6, max_col=6):  # text 컬럼
                        for cell in row:
                            if cell.value and '\n' in str(cell.value):
                                cell.alignment = openpyxl.styles.Alignment(wrap_text=True)
    
    print(f"변환 완료! 엑셀 파일이 생성되었습니다: {output_file_path}")
    return output_file_path

def convert_folder_to_excel(folder_path: str, output_file_path: str = None):
    """폴더 내의 모든 JSON 파일을 하나의 엑셀 파일로 변환합니다."""
    
    # JSON 파일들 찾기
    json_files = glob.glob(os.path.join(folder_path, "*.json"))
    json_files.sort()  # 파일명 순으로 정렬
    
    if not json_files:
        print(f"오류: {folder_path} 폴더에서 JSON 파일을 찾을 수 없습니다.")
        return None
    
    print(f"총 {len(json_files)}개의 JSON 파일을 찾았습니다.")
    
    # 출력 파일명 설정
    if not output_file_path:
        folder_name = os.path.basename(folder_path)
        output_file_path = f"{folder_name}_converted.xlsx"
    
    print(f"엑셀 파일로 변환 중: {output_file_path}")
    
    # 모든 메시지 데이터 수집
    all_messages = []
    
    for json_file in json_files:
        messages_df = process_single_json_file(json_file)
        if not messages_df.empty:
            all_messages.append(messages_df)
    
    if not all_messages:
        print("처리할 메시지가 없습니다.")
        return None
    
    # 모든 메시지 데이터 합치기
    combined_messages = pd.concat(all_messages, ignore_index=True)
    
    # 날짜순으로 정렬
    combined_messages = combined_messages.sort_values('datetime')
    
    # ExcelWriter 객체 생성
    with pd.ExcelWriter(output_file_path, engine='openpyxl') as writer:
        
        # 전체 메시지를 하나의 시트로 저장
        print("전체 메시지를 엑셀로 저장 중...")
        combined_messages.to_excel(writer, sheet_name='All_Messages', index=False)
        
        # 줄바꿈 표시를 위한 셀 서식 설정
        worksheet = writer.sheets['All_Messages']
        for row in worksheet.iter_rows(min_row=2, min_col=6, max_col=6):  # text 컬럼
            for cell in row:
                if cell.value and '\n' in str(cell.value):
                    cell.alignment = openpyxl.styles.Alignment(wrap_text=True)
        
        # 날짜별로 시트 분리
        print("날짜별 시트 생성 중...")
        combined_messages['date'] = pd.to_datetime(combined_messages['datetime']).dt.date
        
        for date, group in combined_messages.groupby('date'):
            date_str = date.strftime('%Y-%m-%d')
            sheet_name = f"Date_{date_str}"
            group_sorted = group.sort_values('datetime')
            group_sorted.to_excel(writer, sheet_name=sheet_name, index=False)
            
            # 줄바꿈 표시를 위한 셀 서식 설정
            worksheet = writer.sheets[sheet_name]
            for row in worksheet.iter_rows(min_row=2, min_col=6, max_col=6):  # text 컬럼
                for cell in row:
                    if cell.value and '\n' in str(cell.value):
                        cell.alignment = openpyxl.styles.Alignment(wrap_text=True)
        
        # 통계 정보 시트 생성
        print("통계 정보 생성 중...")
        stats_data = []
        
        # 날짜별 메시지 수
        daily_stats = combined_messages.groupby('date').size().reset_index(name='message_count')
        for _, row in daily_stats.iterrows():
            stats_data.append({
                '통계_유형': '날짜별_메시지수',
                '날짜': row['date'].strftime('%Y-%m-%d'),
                '값': row['message_count']
            })
        
        # 사용자별 메시지 수
        user_stats = combined_messages.groupby('username').size().reset_index(name='message_count')
        for _, row in user_stats.iterrows():
            stats_data.append({
                '통계_유형': '사용자별_메시지수',
                '사용자': row['username'],
                '값': row['message_count']
            })
        
        # 파일별 메시지 수
        file_stats = combined_messages.groupby('source_file').size().reset_index(name='message_count')
        for _, row in file_stats.iterrows():
            stats_data.append({
                '통계_유형': '파일별_메시지수',
                '파일명': row['source_file'],
                '값': row['message_count']
            })
        
        stats_df = pd.DataFrame(stats_data)
        stats_df.to_excel(writer, sheet_name='Statistics', index=False)
    
    print(f"변환 완료! 엑셀 파일이 생성되었습니다: {output_file_path}")
    print(f"총 {len(combined_messages)}개의 메시지가 처리되었습니다.")
    return output_file_path

def main():
    parser = argparse.ArgumentParser(description='슬랙 채널 백업 JSON 파일을 엑셀 파일로 변환합니다.')
    parser.add_argument('input_path', help='변환할 JSON 파일 경로 또는 폴더 경로')
    parser.add_argument('-o', '--output', help='출력 엑셀 파일 경로 (기본값: 입력파일명_converted.xlsx)')
    parser.add_argument('--folder', action='store_true', help='입력 경로를 폴더로 처리 (폴더 내 모든 JSON 파일을 통합)')
    
    args = parser.parse_args()
    
    if not os.path.exists(args.input_path):
        print(f"오류: 파일 또는 폴더를 찾을 수 없습니다: {args.input_path}")
        return
    
    try:
        if args.folder or os.path.isdir(args.input_path):
            # 폴더 처리
            output_file = convert_folder_to_excel(args.input_path, args.output)
        else:
            # 단일 파일 처리
            output_file = convert_slack_json_to_excel(args.input_path, args.output)
        
        if output_file:
            print(f"\n✅ 변환 성공!")
            print(f"📁 출력 파일: {output_file}")
        else:
            print("❌ 변환에 실패했습니다.")
            
    except Exception as e:
        print(f"❌ 변환 중 오류가 발생했습니다: {str(e)}")

if __name__ == "__main__":
    main()
